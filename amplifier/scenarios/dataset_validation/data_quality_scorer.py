from __future__ import annotations

import json
import sys
from collections.abc import Mapping
from dataclasses import asdict
from dataclasses import dataclass
from pathlib import Path

import click

# Support running as a script: python amplifier/scenarios/dataset_validation/data_quality_scorer.py
if __package__ is None:  # pragma: no cover - runtime path fixup
    project_root = Path(__file__).resolve().parents[3]
    sys.path.append(str(project_root))
    __package__ = "amplifier.scenarios.dataset_validation"

from amplifier.scenarios.dataset_validation.completeness_checker import CompletenessResult
from amplifier.scenarios.dataset_validation.completeness_checker import compute_completeness
from amplifier.scenarios.dataset_validation.format_ranker import FormatScore
from amplifier.scenarios.dataset_validation.format_ranker import score_format
from amplifier.scenarios.dataset_validation.metadata_scorer import MetadataScore
from amplifier.scenarios.dataset_validation.metadata_scorer import score_metadata
from amplifier.scenarios.dataset_validation.models import ColumnStats
from amplifier.scenarios.dataset_validation.models import DatasetProfile
from amplifier.scenarios.dataset_validation.temporal_analyzer import TemporalCoverageResult
from amplifier.scenarios.dataset_validation.temporal_analyzer import analyze_temporal_coverage
from amplifier.scenarios.dataset_validation.temporal_analyzer import parse_date

MISSING_TOKENS = {"", "na", "n/a", "null", "none", "nan"}


@dataclass
class Scorecard:
    composite: float
    completeness: CompletenessResult
    temporal: TemporalCoverageResult
    format: FormatScore
    metadata: MetadataScore


def detect_format(path: Path, explicit_format: str | None) -> str:
    """Detect dataset format from explicit input or file extension."""
    if explicit_format:
        return explicit_format.lower()

    suffix = path.suffix.lower()
    if suffix in {".csv"}:
        return "csv"
    if suffix in {".json"}:
        return "json"
    if suffix in {".xls"}:
        return "xls"
    if suffix in {".xlsx"}:
        return "xlsx"

    raise click.UsageError("Could not detect format. Provide --format (csv|json|xls|xlsx).")


def normalize_cell(value: object) -> object | None:
    """Normalize cell value and treat common missing tokens as None."""
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip()
        if text.lower() in MISSING_TOKENS:
            return None
        return text

    return value


def process_row(
    row: Mapping[str, object],
    stats: dict[str, ColumnStats],
    rows_seen: int,
    date_hints: set[str] | None = None,
) -> None:
    """Update column statistics for a single row."""
    known_columns = set(stats.keys())
    new_columns = set(row.keys()) - known_columns

    # Backfill totals for new columns so missing values in earlier rows are counted
    for col in new_columns:
        stats[col] = ColumnStats(name=col, total=rows_seen)

    all_columns = set(stats.keys())
    date_hints = {h.lower() for h in date_hints} if date_hints else None

    for column in all_columns:
        raw_value = row.get(column)
        value = normalize_cell(raw_value)

        parsed_date = None
        should_parse_date = date_hints is None or column.lower() in date_hints
        if should_parse_date and value is not None:
            parsed_date = parse_date(value)

        stats[column].register(value=value, parsed_date=parsed_date)


def scan_csv(path: Path, date_hints: set[str] | None = None) -> DatasetProfile:
    import csv

    stats: dict[str, ColumnStats] = {}
    rows = 0
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            for field in reader.fieldnames:
                stats[field] = ColumnStats(name=field)

        for row in reader:
            rows += 1
            process_row(row, stats, rows_seen=rows, date_hints=date_hints)

    return DatasetProfile(rows=rows, columns=list(stats.keys()), column_stats=stats)


def scan_json(path: Path, date_hints: set[str] | None = None) -> DatasetProfile:
    stats: dict[str, ColumnStats] = {}
    rows = 0
    data = json.loads(path.read_text())

    if isinstance(data, dict):
        # Attempt to find record-like content
        if "data" in data and isinstance(data["data"], list):
            records = data["data"]
        else:
            raise click.UsageError("JSON must be a list of records or contain a 'data' list.")
    elif isinstance(data, list):
        records = data
    else:
        raise click.UsageError("JSON must be a list of records.")

    for record in records:
        if not isinstance(record, Mapping):
            continue
        rows += 1
        process_row(record, stats, rows_seen=rows, date_hints=date_hints)

    return DatasetProfile(rows=rows, columns=list(stats.keys()), column_stats=stats)


def scan_xlsx(path: Path, date_hints: set[str] | None = None) -> DatasetProfile:
    import openpyxl

    stats: dict[str, ColumnStats] = {}
    rows = 0
    workbook = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active

    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return DatasetProfile(rows=0, columns=[], column_stats=stats)

    header_names = [str(h).strip() for h in headers if h is not None and str(h).strip()]
    for name in header_names:
        stats[name] = ColumnStats(name=name)

    for row_values in iterator:
        rows += 1
        row_dict = {header_names[idx]: row_values[idx] for idx in range(min(len(header_names), len(row_values)))}
        process_row(row_dict, stats, rows_seen=rows, date_hints=date_hints)

    return DatasetProfile(rows=rows, columns=list(stats.keys()), column_stats=stats)


def scan_xls(path: Path, date_hints: set[str] | None = None) -> DatasetProfile:
    import xlrd

    stats: dict[str, ColumnStats] = {}
    rows = 0
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)

    if sheet.nrows == 0:
        return DatasetProfile(rows=0, columns=[], column_stats=stats)

    headers = []
    for cell in sheet.row(0):
        header = str(cell.value).strip()
        if header:
            headers.append(header)
            stats[header] = ColumnStats(name=header)

    for row_idx in range(1, sheet.nrows):
        rows += 1
        row_dict: dict[str, object] = {}
        for col_idx, header in enumerate(headers):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    value = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
                except Exception:
                    value = cell.value
            row_dict[header] = value

        process_row(row_dict, stats, rows_seen=rows, date_hints=date_hints)

    return DatasetProfile(rows=rows, columns=list(stats.keys()), column_stats=stats)


def scan_dataset(path: Path, dataset_format: str, date_hints: set[str] | None) -> DatasetProfile:
    """Scan dataset into profile based on format."""
    loader_map = {
        "csv": scan_csv,
        "json": scan_json,
        "xls": scan_xls,
        "xlsx": scan_xlsx,
    }
    if dataset_format not in loader_map:
        raise click.UsageError(f"Unsupported format '{dataset_format}'. Supported: csv, json, xls, xlsx.")

    return loader_map[dataset_format](path, date_hints=date_hints)


def build_scorecard(
    profile: DatasetProfile,
    dataset_format: str,
    description: str | None,
    tags: list[str] | None,
) -> Scorecard:
    completeness = compute_completeness(profile.column_stats, row_count=profile.rows)
    temporal = analyze_temporal_coverage(profile.column_stats, row_count=profile.rows)
    fmt_score = score_format(dataset_format)
    metadata = score_metadata(description=description, tags=tags)

    composite = round(
        completeness.score * 0.3 + temporal.score * 0.2 + fmt_score.score * 0.25 + metadata.score * 0.25,
        4,
    )

    return Scorecard(
        composite=composite,
        completeness=completeness,
        temporal=temporal,
        format=fmt_score,
        metadata=metadata,
    )


def render_report(
    scorecard: Scorecard,
    profile: DatasetProfile,
    dataset_id: str | None,
    dataset_format: str,
    output_path: Path | None,
) -> None:
    result = {
        "dataset_id": dataset_id,
        "format": dataset_format,
        "rows": profile.rows,
        "columns": profile.columns,
        "scores": {
            "composite": scorecard.composite,
            "completeness": asdict(scorecard.completeness),
            "temporal": {
                **asdict(scorecard.temporal),
                "earliest": scorecard.temporal.earliest.isoformat() if scorecard.temporal.earliest else None,
                "latest": scorecard.temporal.latest.isoformat() if scorecard.temporal.latest else None,
            },
            "format": asdict(scorecard.format),
            "metadata": asdict(scorecard.metadata),
        },
    }

    if output_path:
        output_path.write_text(json.dumps(result, indent=2, ensure_ascii=False))
        click.echo(f"Wrote quality report to {output_path}")
    else:
        click.echo(json.dumps(result, indent=2, ensure_ascii=False))


@click.command()
@click.option("--input", "-i", "input_path", required=True, type=click.Path(path_type=Path, exists=True))
@click.option("--format", "dataset_format", type=click.Choice(["csv", "json", "xls", "xlsx"], case_sensitive=False))
@click.option("--dataset-id", type=str, help="Optional dataset identifier used in the output report")
@click.option("--description", type=str, help="Dataset description used for metadata scoring")
@click.option("--tag", "tags", multiple=True, help="Metadata tag (can be repeated)")
@click.option("--date-column", "date_columns", multiple=True, help="Optional hint for date column names")
@click.option("--output", "-o", "output_path", type=click.Path(path_type=Path), help="Path to write JSON report")
def main(
    input_path: Path,
    dataset_format: str | None,
    dataset_id: str | None,
    description: str | None,
    tags: tuple[str, ...],
    date_columns: tuple[str, ...],
    output_path: Path | None,
) -> None:
    """Run dataset quality scoring pipeline."""
    resolved_format = detect_format(input_path, dataset_format)
    date_hints = set(date_columns) if date_columns else None

    profile = scan_dataset(input_path, resolved_format, date_hints=date_hints)
    scorecard = build_scorecard(profile, resolved_format, description=description, tags=list(tags))
    render_report(scorecard, profile, dataset_id, resolved_format, output_path)


if __name__ == "__main__":
    main()
